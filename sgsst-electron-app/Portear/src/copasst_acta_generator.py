import json
import sys
import os
from openpyxl import load_workbook

def log_message(message, level="INFO"):
    """Generates a JSON-formatted log message."""
    print(json.dumps({"type": "log", "message": message, "level": level}))

def main(temp_data_path, output_path):
    """
    Loads an Excel template, applies changes from a JSON file,
    and saves the result to a new file, preserving styles.
    """
    try:
        log_message("Iniciando generación de acta de COPASST...")

        # 1. Cargar datos del JSON temporal
        with open(temp_data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        changes = data.get('changes', [])
        log_message(f"Cambios a aplicar: {len(changes)}")

        # 2. Ruta a la plantilla
        template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'utils', 'ACT-FO-029 Acta de Reunión Copasst Enero.xlsx')
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en: {template_path}")

        log_message(f"Cargando plantilla desde: {template_path}")

        # 3. Cargar la plantilla con openpyxl
        wb = load_workbook(template_path)
        ws = wb.active

        # 4. Aplicar cambios a las celdas
        for change in changes:
            # openpyxl es 1-indexed, la UI (y JS) es 0-indexed. Se suma 1.
            row = change.get('row') + 1
            col = change.get('col') + 1
            value = change.get('value')
            
            if row is not None and col is not None:
                ws.cell(row=row, column=col, value=value)
        
        log_message("Todos los cambios han sido aplicados a la plantilla en memoria.")

        # 5. Guardar el nuevo archivo en la ruta de salida especificada por el usuario
        wb.save(output_path)
        log_message(f"Acta generada exitosamente en: {output_path}")

        # 6. Devolver resultado de éxito
        print(json.dumps({
            "type": "result",
            "payload": {
                "success": True,
                "documentPath": output_path
            }
        }))

    except Exception as e:
        log_message(f"Error crítico generando el acta: {str(e)}", "ERROR")
        print(json.dumps({
            "type": "result",
            "payload": {
                "success": False,
                "error": str(e)
            }
        }))
        sys.exit(1)

if __name__ == "__main__":
    # El script espera 2 argumentos: la ruta al JSON temporal y la ruta de salida final
    if len(sys.argv) < 3:
        log_message("Uso: python copasst_acta_generator.py <ruta_json_temporal> <ruta_salida_excel>", "ERROR")
        sys.exit(1)
    
    temp_json_path = sys.argv[1]
    final_output_path = sys.argv[2]
    main(temp_json_path, final_output_path)